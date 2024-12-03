from django.contrib.admin.models import CHANGE, LogEntry, ADDITION, DELETION
from django.contrib.contenttypes.models import ContentType
from openpyxl import Workbook
from openpyxl import drawing
from openpyxl.utils import get_column_letter
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.styles import Alignment,Font, Border, Side, PatternFill
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU

def historial_crear_proveedor(user,proveedor):
    LogEntry.objects.log_action(
    user_id         = user.pk, 
    content_type_id = ContentType.objects.get_for_model(proveedor).pk,
    object_id       = proveedor.pk,
    object_repr     = "Proveedor {}".format(proveedor.empresa_nombre),
    action_flag     = ADDITION,
    change_message = "El proveedor {} ha sido creado por el usuario {}".format(proveedor.empresa_nombre,user)
    )


def historial_eliminar_proveedor(user,proveedor):
    LogEntry.objects.log_action(
    user_id         = user.pk, 
    content_type_id = ContentType.objects.get_for_model(proveedor).pk,
    object_id       = proveedor.pk,
    object_repr     = "Proveedor {}".format(proveedor.empresa_nombre),
    action_flag     = DELETION,
    change_message = "El proveedor {} ha sido eliminado por el usuario {}".format(proveedor.empresa_nombre,user)
    )

def historial_actualizar_proveedor(user,proveedor):
    LogEntry.objects.log_action(
    user_id         = user.pk, 
    content_type_id = ContentType.objects.get_for_model(proveedor).pk,
    object_id       = proveedor.pk,
    object_repr     = "Proveedor {}".format(proveedor.empresa_nombre),
    action_flag     = CHANGE,
    change_message = "El proveedor {} ha sido actualizado por el usuario {}".format(proveedor.empresa_nombre,user)
    )

def generar_excel(asociaciones, request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Insumos'
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    encabezados = ['Item','Categoria MP','Materia Prima','Unidad','Precio','Proveedor']
    for i in range(len(encabezados)):
        ws.cell(row=1,column=i+1,value = encabezados[i]).border = thin_border
    comienzo = 2
    for a in asociaciones:
        ws.cell(row=comienzo,column=1,value=a.insumo.id)
        ws.cell(row=comienzo,column=2,value="")
        ws.cell(row=comienzo,column=3,value=a.insumo.nombre)
        ws.cell(row=comienzo,column=4,value=a.insumo.unidad)
        print(a.moneda.valor)
        try:
            ws.cell(row=comienzo,column=5,value=a.precio/a.formato * a.moneda.valor)
        except:
            ws.cell(row=comienzo,column=5,value=a.precio * a.moneda.valor)
        ws.cell(row=comienzo,column=6,value=a.proveedor.empresa_nombre)
        comienzo+=1
    return wb