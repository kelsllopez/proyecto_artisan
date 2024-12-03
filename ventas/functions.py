from django.conf import settings
from openpyxl import Workbook
from django.urls import reverse
from django.templatetags.static import static
from openpyxl import drawing
from openpyxl.utils import get_column_letter
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from ordendecompra.templatetags.filtros import moneda


def generar_excel(ov, request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'OV'
    ws.page_setup.paperSize = 1
    ws.page_margins.left = 0.23622
    ws.page_margins.right = 0.23622
    ws.page_margins.top = 0.7519685
    ws.page_margins.header = 0.2992126
    ws.page_margins.bottom = 0.7519685
    ws.page_margins.footer = 0.2992126
    font = Font(size=8)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws['A3'].value = 'Cliente'
    ws['C3'].value = ov.cliente.nombre
    ws['A4'].value = 'Rut'
    ws['C4'].value = ov.cliente.rut
    ws['A5'].value = 'Razón Social'
    ws['C5'].value = ov.cliente.razon_social
    ws['A6'].value = 'Direccion'
    ws['C6'].value = ov.local.direccion
    ws['A7'].value = 'Comuna'
    ws['C7'].value = ov.local.comuna
    ws['A8'].value = 'Horario'
    ws['C8'].value = 'Continuado'
    ws['A9'].value = 'Condiciones Pago'
    ws['C9'].value = ov.cliente.pago
    ws['F2'].value = ''
    ws['G2'].value = "${}".format(moneda(ov.total()))
    ws['F3'].value = 'Factura'
    ws['G3'].value = ''
    ws['F4'].value = 'Neto'
    ws['G4'].value = "${}".format(moneda(ov.totalNeto()))
    ws['F5'].value = 'Iva'
    ws['G5'].value = "${}".format(moneda(ov.iva()))
    ws['F6'].value = 'Total'
    ws['G6'].value = "${}".format(moneda(ov.total()))
    ws['F8'] = 'Rut: 76.059.975-1'
    ws['F8'].font = font
    ws['F9'] = 'Elaboradora de Alimentos Gourmet Ltda.'
    ws['F9'].font = font
    ws['F10'] = 'Cuenta Corriente 490370201'
    ws['F10'].font = font
    ws['F11'] = 'Banco de Chile'
    ws['F11'].font = font
    ws['H11'] = 'oc@quesosartisan.cl'
    ws['H11'].font = font

    ws['B14'] = 'Cliente'
    ws['C14'] = 'N° Factura'
    ws['D14'] = 'Producto'
    ws['E14'] = 'Cant Uni OC'
    ws['F14'] = 'Cant Uni FC'
    ws['G14'] = '% Desc.'
    ws['H14'] = 'Precio Neto'
    ws['I14'] = 'Neto FC'
    ws['J14'] = 'SUM de IVA FC'
    ws['K14'] = 'SUM de Monto Bruto FC'

    letras = ['B','C','D','E','F','G','H','I','J','K']
    for l in letras:
        ws['{}14'.format(l)].font = Font(size=8,bold=True)
        ws['{}14'.format(l)].border = Border(bottom=Side(style='thin'))
        ws['{}14'.format(l)].fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type = "solid")
    ws['B15'] = ov.cliente.nombre
    indice = 15
    letras = ['B','C','D','E','F','G','H','I','J','K']
    for producto in ov.ordendeventaproducto_set.all():
        ws['D{}'.format(indice)] = "{} {} {}".format(producto.producto.nombre.replace('Helado','H'), producto.producto.presentacion, producto.producto.unidad.upper())
        ws['E{}'.format(indice)] = producto.cantidad
        ws['H{}'.format(indice)] = "${:10,.2f}".format(producto.precio)
        for l in letras:
            ws['{}{}'.format(l,indice)].font = font
            ws['{}{}'.format(l,indice)].fill = PatternFill(start_color='F3F3F3', end_color='F3F3F3', fill_type = "solid")
            ws['{}{}'.format(l,indice)].border = Border(bottom=Side(style='thin'))
        indice+=1
    estados = ["No pickeado", "Pick Incompleto", "Pick Completo", "Fact", "Entregado"]
    indice = 3
    for e in estados:
        ws['I{}'.format(indice)] = e
        ws['J{}'.format(indice)] = ""
        ws['I{}'.format(indice)].border = thin_border
        ws['J{}'.format(indice)].border = thin_border
        ws['I{}'.format(indice)].font = font
        ws['I{}'.format(indice)].font = font
        indice+=1
    for i in range(2,7):
        letras = ['F','G']
        for l in letras:
            ws['{}{}'.format(l,i)].border = thin_border
            if l == 'F':
                ws['{}{}'.format(l,i)].font = Font(size=8,bold=True)
            else:
                ws['{}{}'.format(l,i)].font = font
    combinar = [('A', 'B'), ('C', 'D')]
    for i in range(3, 10):
        for c in combinar:
            ws.merge_cells('{}{}:{}{}'.format(c[0], i, c[1], i))
            ws['{}{}'.format(c[0], i)].border = thin_border
            ws['{}{}'.format(c[1], i)].border = thin_border
            if c[0] == 'A':
                ws['{}{}'.format(c[0], i)].font = Font(size=8,bold=True)
            else:
                ws['{}{}'.format(c[0], i)].font = font
    return wb

def exportar_excel(fechai,fechaf,ovs,request):
    c2e = cm_to_EMU
    p2e = pixels_to_EMU
    # Calculated number of cells width or height from cm into EMUs
    cellh = lambda x: c2e((x * 49.77)/99)
    cellw = lambda x: c2e((x * (18.65-1.71))/10)
    centrado = Alignment(horizontal='center',vertical='center')
    font = Font(size=16)
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    wb = Workbook()
    ws = wb.active
    ws.title = 'Libro de Ventas'
    encabezados = ['Fecha OC','Código Producto','Producto','Cantidad OC','Cantidad FC','Fecha FC','gr u','u x caja','kg oc','kg fc','precio u','monto neto oc','monto neto fc','iva fc','monto bruto fc','cat producto']
    for i in range(len(encabezados)):
        ws.cell(row=1,column=i+1).value = encabezados[i]
    row = 2
    unidades = {'lt':1000,'gr':1,'ml':1,'cc':1,'kg':1000}
    for ov in ovs:
        productos = ov.ordendeventaproducto_set.all()
        for p in productos:
            if ov.estado == 'Recepción Parcial':
                p.cantidad_fc = p.cantidad_en
            ws.cell(row=row,column=1).value = ov.fecha.strftime('%d/%m/%Y')
            ws.cell(row=row,column=2).value = p.producto.codigo
            ws.cell(row=row,column=3).value = "{} ({} {})".format(p.producto.nombre,p.producto.presentacion,p.producto.unidad)
            ws.cell(row=row,column=4).value = p.cantidad
            ws.cell(row=row,column=5).value = p.cantidad_fc
            if ov.fecha_f is None:
                ws.cell(row=row,column=6).value = 'No'
            else:
                ws.cell(row=row,column=6).value = ov.fecha_f.strftime('%d/%m/%Y')
            ws.cell(row=row,column=7).value = p.producto.presentacion * unidades[p.producto.unidad]
            ws.cell(row=row,column=8).value = p.producto.unidades
            divisor = 1000
            if p.producto.unidad not in ['gr', 'cc', 'ml', 'unidad', 'caja']:
                divisor=1
            ws.cell(row=row,column=9).value = round((p.producto.presentacion * p.cantidad)/divisor,2)
            ws.cell(row=row,column=10).value = round((p.producto.presentacion * p.cantidad_fc)/divisor,2)
            ws.cell(row=row,column=11).value = p.precio
            ws.cell(row=row,column=12).value = p.precio * p.cantidad
            ws.cell(row=row,column=13).value = p.precio * p.cantidad_fc
            ws.cell(row=row,column=14).value = p.precio * p.cantidad_fc * settings.IVA
            ws.cell(row=row,column=15).value = p.precio * p.cantidad_fc * settings.IVA + p.precio * p.cantidad_fc
            ws.cell(row=row,column=16).value = p.producto.rama.nombre
            row+=1
    for i in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(i)].bestFit = True
        ws.column_dimensions[get_column_letter(i)].auto_size = True
    ws.auto_filter.ref = 'A1:C1'
    return wb

