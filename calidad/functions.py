from openpyxl import Workbook
from openpyxl import drawing
from openpyxl.utils import get_column_letter
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.styles import Alignment,Font, Border, Side, PatternFill
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from django.urls import reverse
from django.templatetags.static import static


def generar_excel(fechai,fechaf,registros,request):
    c2e = cm_to_EMU
    p2e = pixels_to_EMU
    # Calculated number of cells width or height from cm into EMUs
    cellh = lambda x: c2e((x * 49.77)/99)
    cellw = lambda x: c2e((x * (18.65-1.71))/10)
    centrado = Alignment(horizontal='center',vertical='center')
    font = Font(size=16)
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    wb = Workbook()
    ws = wb.active
    ws.title = 'Registros'
    img = drawing.image.Image('static/nucleo/img/logo.jpg')
    img.height = 65
    img.width = 65
    img.anchor = 'A1'
    combinar = ['A1:A3','B1:H1','B2:H3','I1:J1','I2:J3']
    centrar = ['B1','I2','I1','B2']
    for c in combinar:
        ws.merge_cells(c)
    ws.cell(row=1,column=2,value = 'Código: R-AC-02').font = font
    for c in centrar:
        ws[c].alignment = centrado
    ws['B2'].value = 'REGISTRO DE LIMPIEZA Y SANITIZACIÓN DE EQUIPOS Y ÁREAS'
    ws['I1'].value = 'Rev.01'
    ws['I1'].font = font
    ws['I2'].value = 'Fecha: {} a {}'.format(fechai,fechaf)
    ws['B2'].font = Font(size=16,bold=True)
    ws.add_image(img)
    for i in range(1,4,1):
        for j in range(1,11,1):
            ws.cell(row=i,column=j).border = thin_border
    encabezados = ['Fecha/Hora','Área','Operador','Máquina','Artículos de limpieza','Observaciones','Firma operador','Acciones correctivas','Estado de la operación','Firma Supervisor']
    for i in range(len(encabezados)):
        ws.cell(row=4,column=i+1,value = encabezados[i]).border = thin_border
    comienzo = 5
    colores = {'Ejecutado':'FFC7CE','Pendiente':'fdffc7','Aprobado':'d0ffc7'}
    for r in registros:
        operador = r.encargado
        supervisor = r.revisado
        utensilios = ""
        for u in r.utensilios.all():
            utensilios+=u.nombre+", "
        utensilios = utensilios[0:len(utensilios)-2]
        ws.cell(row=comienzo,column=1,value=r.created)
        if operador:
            ws.cell(row=comienzo,column=3,value=operador.first_name + " " + operador.last_name)
        else:
            ws.cell(row=comienzo,column=3,value="Sin Operador")
        if r.equipo.area:
            ws.cell(row=comienzo,column=2,value=r.equipo.area.nombre)
        else:
            ws.cell(row=comienzo,column=2,value="Sin área")
        ws.cell(row=comienzo,column=4,value=r.equipo.nombre)
        ws.cell(row=comienzo,column=5,value=utensilios)
        ws.cell(row=comienzo,column=6,value=r.observacion)
        #ver si el operario tiene firma digital
        if operador:
            if operador.perfil.firma_digital:
                img = drawing.image.Image('media/{}'.format(operador.perfil.firma_digital))
                img.height = 30
                img.width = 150
                coloffset = cellw(0.25)
                rowoffset = cellh(0.5)
                marker = AnchorMarker(col=6, colOff=coloffset, row=comienzo-1, rowOff=rowoffset)
                h, w = img.height, img.width
                size = (XDRPositiveSize2D(p2e(w), p2e(h)))
                img.anchor = OneCellAnchor(_from=marker, ext=size)
                ws.add_image(img)
            else:
                ws.cell(row=comienzo,column=7,value="sin firma")
        else:
            ws.cell(row=comienzo,column=7,value="sin firma")
        ws.cell(row=comienzo,column=8,value=r.accion_correctiva)
        ws.cell(row=comienzo,column=9,value=r.estado).fill = PatternFill(start_color=colores[r.estado], end_color=colores[r.estado], fill_type = "solid")

        #si tiene supervisor
        if supervisor:
            if supervisor.perfil.firma_digital:
                img = drawing.image.Image('media/{}'.format(supervisor.perfil.firma_digital))
                img.height = 30
                img.width = 150
                coloffset = cellw(0.25)
                rowoffset = cellh(0.5)
                marker = AnchorMarker(col=9, colOff=coloffset, row=comienzo-1, rowOff=rowoffset)
                h, w = img.height, img.width
                size = (XDRPositiveSize2D(p2e(w), p2e(h)))
                img.anchor = OneCellAnchor(_from=marker, ext=size)
                ws.add_image(img)
            else:
                ws.cell(row=comienzo,column=10,value="sin firma")
        else:
            ws.cell(row=comienzo,column=10,value="Sin Supervisor")
        #bordes
        for i in range(1,11,1):
            ws.cell(row=comienzo,column=i).border = thin_border
        comienzo+=1
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].bestFit = True
    for i in range(1,len(registros)+1,1):
        ws.row_dimensions[4+i].height = 35
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['J'].width = 40
    ws1 = wb.create_sheet("Historial")
    encabezados = ['Fecha/Hora','Área','Lugar','Máquina','Operador','Observaciones','Acciones correctivas','Estado de la operación','Supervisor']
    for i in range(len(encabezados)):
        ws1.cell(row=1,column=i+1,value = encabezados[i]).border = thin_border
    pos = 2
    for r in registros:
        historiales = r.registrolimpiezaequipohistorial_set.all()
        operador = r.encargado
        supervisor = r.revisado
        for h in historiales:
            ws1.cell(row=pos,column=1,value=r.created).border = thin_border
            if r.equipo.area:
                ws1.cell(row=pos,column=2,value=r.equipo.area.nombre).border = thin_border
                ws1.cell(row=pos,column=3,value=r.equipo.area.lugar.nombre).border = thin_border
            else:
                ws1.cell(row=pos,column=2,value="Sin área").border = thin_border
                ws1.cell(row=pos,column=3,value="Sin lugar").border = thin_border
            ws1.cell(row=pos,column=4,value=r.equipo.nombre).border = thin_border
            if operador:
                ws1.cell(row=pos,column=5,value=operador.first_name + " " + operador.last_name).border = thin_border
            else:
                ws1.cell(row=pos,column=5,value="Sin Operador").border = thin_border
            ws1.cell(row=pos,column=6,value=h.observacion).border = thin_border
            ws1.cell(row=pos,column=7,value=h.accion_correctiva).border = thin_border
            ws1.cell(row=pos,column=8,value=h.estado).fill = PatternFill(start_color=colores[h.estado], end_color=colores[h.estado], fill_type = "solid")
            if supervisor:
                ws1.cell(row=pos,column=9,value=supervisor.first_name + " " + supervisor.last_name).border = thin_border
            else:
                ws1.cell(row=pos,column=9,value="Sin Operador").border = thin_border
            pos+=1
    for idx, col in enumerate(ws1.columns, 1):
        ws1.column_dimensions[get_column_letter(idx)].bestFit = True
        ws1.column_dimensions[get_column_letter(idx)].auto_size = True
    return wb
    